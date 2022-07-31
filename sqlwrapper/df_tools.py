import pandas as pd
import pathlib
#from typing import Union

def max_len_cols(df_input, method='default', *args,  **kwargs):
    """manages calling the two functions"""
    if method == 'oracle':
        return max_len_cols_oracle(df_input, *args, **kwargs)
    else:
        return max_len_cols_default(df_input, *args, **kwargs)

def max_len_cols_default(df_input) -> dict:
    """returns a dict of max count of each column in pd.DataFrame"""
    max_len = {}
    df_input = df_input.fillna('')
    for col in df_input.columns:
        try:
            max_len[col.upper()] = int(df_input[col].str.len().max())
        except AttributeError as error:
            max_len[col] = df_input[col].dtype
    return max_len

def max_len_cols_oracle(df_input, factor=1) -> dict:
    """returns datatype with ORACLE NUMBER TYPE"""
    from sqlalchemy.dialects.oracle import NUMBER, VARCHAR2, DATE, TIMESTAMP
    
    max_len = {}
    for col in df_input.columns:
        try:
            if df_input[col].dtype.name is 'bool':
                max_len[col] = NUMBER(1,0)
            else:
                max_len[col.upper()] = df_input[col].str.len().max()
        except AttributeError as error:
            if 'date' in col.lower():
                max_len[col.upper()] = 'date'
            print(error)
            print(col)
    #  setting datatypes
    result_dict = {}
    for col, dtype in max_len.items():
        try:
            if 'date' in col.lower():
                result_dict[col] = DATE
            elif type(dtype) == NUMBER:
                result_dict[col.upper()] = dtype
            else:
                result_dict[col] = VARCHAR2(int(int(dtype)* factor))
        except:
            print(col, dtype)
            result_dict[col] = VARCHAR2(100)
    return result_dict


def read_xlsx(path_to_xlsx:pathlib.WindowsPath, 
              return_type='default'): #-> Union(pd.DataFrame, dict):
    from openpyxl import load_workbook 

    #path_example: # Path.cwd() / 'data' /'spreadsheet.xlsx'
    workbook = load_workbook(filename=path_to_xlsx)

    if return_type == 'dfs':
        dict_dfs = {}
        for sheet in workbook.sheetnames:
            dict_dfs[sheet] = sheet_to_df(workbook[sheet])
        return dict_dfs
    else: # else just return the workbook
        print(f'Returning workbook, with sheets: {str(workbook.sheetnames)}.')
        return workbook
    


def sheet_to_df(sheet_input, columns=True):
    """
    * input is an openpyxl sheet
    * output is a pandas dataframe
    """
    data = sheet_input.values
    # get column data
    if columns:
        cols = next(data)
    # get data
    data = list(data)
    df_output = pd.DataFrame(data, columns=cols)
    return df_output
