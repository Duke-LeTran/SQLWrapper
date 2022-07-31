from openpyxl import load_workbook 
import pandas as pd
import pathlib


def read_xlsx(path_to_xlsx:pathlib.Path, 
              return_type='default'): #-> Union(pd.DataFrame, dict):


    #path_example: # Path.cwd() / 'data' /'spreadsheet.xlsx'
    workbook = load_workbook(filename=path_to_xlsx)

    if return_type == 'dfs':
        dict_dfs = {}
        for sheet in workbook.sheetnames:
            dict_dfs[sheet] = sheet_to_df(workbook[sheet])
        return dict_dfs
    else: # else just return the workbook
        print(f'Returning workbook, with sheets: {str(workbook.sheetnames)}.')
        return workbook
    

def sheet_to_df(sheet_input, columns=True):
    """
    * input is an openpyxl sheet
    * output is a pandas dataframe
    """
    data = sheet_input.values
    # get column data
    if columns:
        cols = next(data)
    # get data
    data = list(data)
    df_output = pd.DataFrame(data, columns=cols)
    return df_output
